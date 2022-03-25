import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.xml.constants import MIN_COLUMN

wb = xl.load_workbook("C:/Users/heman/Desktop/Programming/Python/automation/transactions.xlsx")

sheet = wb["Sheet1"]

cell = sheet['a1']
cell = sheet.cell(1, 1)
# Line no 7 and 8 does the same thing. Any one can be used...

print(cell)
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)

for row in range(2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   corrected_price = cell.value * 0.9
   corrected_price_cell = sheet.cell(row, 4)
   corrected_price_cell.value = corrected_price
   print(corrected_price_cell.value)


value = Reference(sheet, min_row=2, max_row = sheet.max_row, min_col = 4, max_col = 4)
chart = BarChart()
chart.add_data(value)
sheet.add_chart(chart, 'a6')


wb.save('transactions2.xlsx')
